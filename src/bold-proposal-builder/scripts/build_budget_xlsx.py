"""
build_budget_xlsx.py v4.0

Produces Bold's canonical event budget XLSX from a structured budget.json.

LAYOUT (13 columns, RTL, single sheet "טמפלט ריק"):

  Columns A-F: supplier side (right side in RTL view)
    A  ספק                  vendor name
    B  חשבונית              invoice reference
    C  הוצאה בפועל           actual spend (filled post-event)
    D  תנאי תשלום           payment terms
    E  מחיר ליחידה לספק       supplier unit cost
    F  סה"כ עלות             total cost = qty * unit_cost

  Columns G-K: client-facing (middle in RTL view)
    G  קטגוריה               canonical category (one of 6)
    H  תיאור                  service description
    I  כמות                   units (qty)
    J  מחיר ליחידה             client unit price = unit_cost * 1.15  (HIDDEN markup)
    K  סה"כ                   total charge = qty * unit_price

  Columns L-M: profit (left side in RTL view, gray fill)
    L  רווח                  profitability = total_charge - total_cost
    M  אחוז רווח              margin = profitability / total_charge

MARGIN MODEL (v4.0 dual-layer, replaces v3.0 single-layer):

  Layer 1: Per-line embedded markup of 15%
    client unit_price = supplier unit_cost * 1.15
    This is invisible to the client (no column or label says "15%").
    It manifests only as the gap between Bold's cost (col F) and
    the client's charge (col K), visible in cols L-M (gray profit area).

  Layer 2: Visible production fee of 15%
    A dedicated row "דמי ארגון והפקה" appears after the main section
    and again after the options/conditional section. It equals 15% of
    the section's client subtotal (sum of col K within the section).
    This is shown explicitly to the client.

  Effective Bold margin per shekel charged: ~32% (1.15 * 1.15 = 1.3225)

SECTIONS:

  Main section: lines grouped under the 6 canonical category markers
    (כללי, תקשור מקדים, מיתוג ושילוט, טכני, כח אדם ולוגיסטיקה, שונות)
    Followed by main section subtotal, production fee, grand total.

  Options section: conditional items, flat under a single כללי marker.
    Followed by options subtotal, options production fee, options grand total.

FORMULAS:

  All derivations (total_cost, unit_price, total_charge, profitability,
  margin, subtotals, production fees, grand totals) are written as
  Excel formulas, NOT static values. This lets Hemi edit supplier prices
  in col E and watch every downstream cell recalculate.

USAGE:

  python build_budget_xlsx.py budget.json output.xlsx

INPUT SCHEMA (budget.json):

  {
    "meta": {
      "proposal_name": "...",
      "client": "...",
      "guest_count": 200,
      "currency": "ILS",
      "valid_until": "YYYY-MM-DD"
    },
    "categories": [
      {
        "name": "טכני",          // must be one of the 6 canonicals
        "items": [
          {
            "description": "מערכת הגברה מלאה",
            "qty": 1,
            "unit": "אירוע",
            "unit_cost": 8500,    // supplier price per unit
            "vendor_name": "ABC הגברה",  // optional, supplier side
            "payment_terms": "שוטף+30",  // optional
            "source_deliverable": "operations/logistics.md sec 3"
          }
        ]
      }
    ],
    "conditional_items": [
      {
        "description": "תאורת חצר נוספת",
        "qty": 1,
        "unit": "אירוע",
        "unit_cost": 3200,
        "vendor_name": "...",
        "trigger_condition": "outdoor extension"
      }
    ]
  }

Note: client unit_price is NEVER specified in budget.json. The script
derives it from unit_cost * 1.15. This is the load-bearing principle:
the per-line markup is a script-internal calculation, never a manual
field that could leak to the client.
"""

import json
import sys
from copy import copy
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet


# Bold's six canonical budget categories (in order).
CANONICAL_CATEGORIES = [
    "כללי",
    "תקשור מקדים",
    "מיתוג ושילוט",
    "טכני",
    "כח אדם ולוגיסטיקה",
    "שונות",
]

# Markup rates (both 15%, applied in different layers).
PER_LINE_MARKUP = 0.15           # embedded in unit_price, invisible to client
PRODUCTION_FEE_RATE = 0.15       # visible "דמי ארגון והפקה" line

# Style constants.
SHEET_NAME = "טמפלט ריק"
HEADER_FILL = PatternFill(start_color="2C2C2C", end_color="2C2C2C", fill_type="solid")
HEADER_FONT = Font(name="Arial", size=11, bold=True, color="FFFFFF")
CATEGORY_FILL = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
CATEGORY_FONT = Font(name="Arial", size=11, bold=True, color="000000")
GRAY_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
SUBTOTAL_FILL = PatternFill(start_color="BFBFBF", end_color="BFBFBF", fill_type="solid")
SUBTOTAL_FONT = Font(name="Arial", size=11, bold=True, color="000000")
GRAND_TOTAL_FILL = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
GRAND_TOTAL_FONT = Font(name="Arial", size=12, bold=True, color="FFFFFF")
THIN_BORDER = Border(
    left=Side(border_style="thin", color="999999"),
    right=Side(border_style="thin", color="999999"),
    top=Side(border_style="thin", color="999999"),
    bottom=Side(border_style="thin", color="999999"),
)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True, readingOrder=2)
RIGHT = Alignment(horizontal="right", vertical="center", wrap_text=True, readingOrder=2)

# Column letter constants for readability.
COL_VENDOR = "A"
COL_INVOICE = "B"
COL_ACTUAL = "C"
COL_TERMS = "D"
COL_SUP_UNIT_COST = "E"
COL_TOTAL_COST = "F"
COL_CATEGORY = "G"
COL_DESCRIPTION = "H"
COL_QTY = "I"
COL_UNIT_PRICE = "J"
COL_TOTAL_CHARGE = "K"
COL_PROFIT = "L"
COL_MARGIN = "M"

HEADERS = {
    COL_VENDOR: "ספק",
    COL_INVOICE: "חשבונית",
    COL_ACTUAL: "הוצאה בפועל",
    COL_TERMS: "תנאי תשלום",
    COL_SUP_UNIT_COST: "מחיר ליחידה לספק",
    COL_TOTAL_COST: 'סה"כ עלות',
    COL_CATEGORY: "קטגוריה",
    COL_DESCRIPTION: "תיאור",
    COL_QTY: "כמות",
    COL_UNIT_PRICE: "מחיר ליחידה",
    COL_TOTAL_CHARGE: 'סה"כ',
    COL_PROFIT: "רווח",
    COL_MARGIN: "אחוז רווח",
}

COLUMN_WIDTHS = {
    COL_VENDOR: 18,
    COL_INVOICE: 12,
    COL_ACTUAL: 14,
    COL_TERMS: 14,
    COL_SUP_UNIT_COST: 14,
    COL_TOTAL_COST: 14,
    COL_CATEGORY: 18,
    COL_DESCRIPTION: 36,
    COL_QTY: 8,
    COL_UNIT_PRICE: 14,
    COL_TOTAL_CHARGE: 14,
    COL_PROFIT: 14,
    COL_MARGIN: 11,
}


def normalize_category(raw_name):
    """Return the canonical category name, or warn and return 'שונות'."""
    if raw_name in CANONICAL_CATEGORIES:
        return raw_name
    print(
        f"  [WARN] non-canonical category '{raw_name}', rolling into שונות",
        file=sys.stderr,
    )
    return "שונות"


def group_items_by_canonical(categories):
    """Take input categories and bucket all their items under the 6 canonicals."""
    buckets = {name: [] for name in CANONICAL_CATEGORIES}
    for cat in categories:
        canonical = normalize_category(cat.get("name", ""))
        for item in cat.get("items", []):
            buckets[canonical].append(item)
    return buckets


def write_header_row(ws):
    """Header row at row 1."""
    for col_letter, label in HEADERS.items():
        cell = ws[f"{col_letter}1"]
        cell.value = label
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = THIN_BORDER
    ws.row_dimensions[1].height = 36


def write_category_marker(ws, row, category_name):
    """A full-width band labelled with the category name."""
    cell = ws.cell(row=row, column=8)  # column H (תיאור) is the visual anchor
    cell.value = category_name
    cell.font = CATEGORY_FONT
    cell.alignment = CENTER
    for col_idx in range(1, 14):
        c = ws.cell(row=row, column=col_idx)
        c.fill = CATEGORY_FILL
        c.border = THIN_BORDER
    ws.row_dimensions[row].height = 28


def write_line_item(ws, row, item, category_name):
    """A single line. Formulas applied to derived columns."""
    qty = item.get("qty", 1)
    unit_cost = item.get("unit_cost")
    description = item.get("description", "")
    vendor = item.get("vendor_name", "")
    payment_terms = item.get("payment_terms", "")

    # Supplier side (right).
    ws[f"{COL_VENDOR}{row}"] = vendor
    ws[f"{COL_INVOICE}{row}"] = ""
    ws[f"{COL_ACTUAL}{row}"] = ""
    ws[f"{COL_TERMS}{row}"] = payment_terms
    ws[f"{COL_SUP_UNIT_COST}{row}"] = unit_cost if unit_cost is not None else None
    # Total cost (col F) = qty * supplier unit cost.
    ws[f"{COL_TOTAL_COST}{row}"] = f"={COL_QTY}{row}*{COL_SUP_UNIT_COST}{row}"

    # Client-facing (middle).
    ws[f"{COL_CATEGORY}{row}"] = category_name
    ws[f"{COL_DESCRIPTION}{row}"] = description
    ws[f"{COL_QTY}{row}"] = qty
    # Unit price (col J) = supplier unit cost * 1.15. The 15% never appears as
    # a label or a column header - only as a hardcoded multiplier in this formula.
    ws[f"{COL_UNIT_PRICE}{row}"] = f"={COL_SUP_UNIT_COST}{row}*{1 + PER_LINE_MARKUP}"
    # Total charge (col K) = qty * unit price.
    ws[f"{COL_TOTAL_CHARGE}{row}"] = f"={COL_QTY}{row}*{COL_UNIT_PRICE}{row}"

    # Profit (left, gray).
    ws[f"{COL_PROFIT}{row}"] = f"={COL_TOTAL_CHARGE}{row}-{COL_TOTAL_COST}{row}"
    ws[f"{COL_MARGIN}{row}"] = (
        f"=IF({COL_TOTAL_CHARGE}{row}=0,0,{COL_PROFIT}{row}/{COL_TOTAL_CHARGE}{row})"
    )

    # Format.
    for col_letter in HEADERS:
        c = ws[f"{col_letter}{row}"]
        c.alignment = CENTER if col_letter in (COL_QTY, COL_MARGIN) else RIGHT
        c.border = THIN_BORDER
        if col_letter in (COL_PROFIT, COL_MARGIN):
            c.fill = GRAY_FILL
        if col_letter == COL_MARGIN:
            c.number_format = "0.0%"
        elif col_letter in (
            COL_SUP_UNIT_COST, COL_TOTAL_COST, COL_UNIT_PRICE,
            COL_TOTAL_CHARGE, COL_PROFIT, COL_ACTUAL,
        ):
            c.number_format = '#,##0" ₪"'

    ws.row_dimensions[row].height = 22


def write_section_subtotal(ws, row, label, first_data_row, last_data_row):
    """Subtotal row summing rows first_data_row..last_data_row in K, F, L."""
    ws.cell(row=row, column=8).value = label  # col H = תיאור slot
    for col_idx in range(1, 14):
        c = ws.cell(row=row, column=col_idx)
        c.fill = SUBTOTAL_FILL
        c.font = SUBTOTAL_FONT
        c.alignment = CENTER if col_idx in (9, 13) else RIGHT
        c.border = THIN_BORDER

    ws[f"{COL_TOTAL_COST}{row}"] = f"=SUM({COL_TOTAL_COST}{first_data_row}:{COL_TOTAL_COST}{last_data_row})"
    ws[f"{COL_TOTAL_CHARGE}{row}"] = f"=SUM({COL_TOTAL_CHARGE}{first_data_row}:{COL_TOTAL_CHARGE}{last_data_row})"
    ws[f"{COL_PROFIT}{row}"] = f"=SUM({COL_PROFIT}{first_data_row}:{COL_PROFIT}{last_data_row})"
    ws[f"{COL_MARGIN}{row}"] = (
        f"=IF({COL_TOTAL_CHARGE}{row}=0,0,{COL_PROFIT}{row}/{COL_TOTAL_CHARGE}{row})"
    )

    ws[f"{COL_TOTAL_COST}{row}"].number_format = '#,##0" ₪"'
    ws[f"{COL_TOTAL_CHARGE}{row}"].number_format = '#,##0" ₪"'
    ws[f"{COL_PROFIT}{row}"].number_format = '#,##0" ₪"'
    ws[f"{COL_MARGIN}{row}"].number_format = "0.0%"
    ws.row_dimensions[row].height = 26


def write_production_fee(ws, row, subtotal_row):
    """The visible 15% production fee line. Client sees this as a labeled item."""
    ws.cell(row=row, column=8).value = "דמי ארגון והפקה"
    ws[f"{COL_TOTAL_CHARGE}{row}"] = f"={COL_TOTAL_CHARGE}{subtotal_row}*{PRODUCTION_FEE_RATE}"
    ws[f"{COL_PROFIT}{row}"] = f"={COL_TOTAL_CHARGE}{row}"  # full production fee is profit
    ws[f"{COL_MARGIN}{row}"] = 1.0  # 100% margin on this line by definition

    for col_idx in range(1, 14):
        c = ws.cell(row=row, column=col_idx)
        c.border = THIN_BORDER
        c.alignment = CENTER if col_idx in (9, 13) else RIGHT
        if col_idx in (12, 13):
            c.fill = GRAY_FILL
    ws[f"{COL_TOTAL_CHARGE}{row}"].number_format = '#,##0" ₪"'
    ws[f"{COL_PROFIT}{row}"].number_format = '#,##0" ₪"'
    ws[f"{COL_MARGIN}{row}"].number_format = "0.0%"
    ws.row_dimensions[row].height = 24


def write_grand_total(ws, row, subtotal_row, fee_row, label='סה"כ חשבונית'):
    """Final total = subtotal + production fee."""
    ws.cell(row=row, column=8).value = label
    ws[f"{COL_TOTAL_COST}{row}"] = f"={COL_TOTAL_COST}{subtotal_row}"
    ws[f"{COL_TOTAL_CHARGE}{row}"] = f"={COL_TOTAL_CHARGE}{subtotal_row}+{COL_TOTAL_CHARGE}{fee_row}"
    ws[f"{COL_PROFIT}{row}"] = f"={COL_TOTAL_CHARGE}{row}-{COL_TOTAL_COST}{row}"
    ws[f"{COL_MARGIN}{row}"] = (
        f"=IF({COL_TOTAL_CHARGE}{row}=0,0,{COL_PROFIT}{row}/{COL_TOTAL_CHARGE}{row})"
    )

    for col_idx in range(1, 14):
        c = ws.cell(row=row, column=col_idx)
        c.fill = GRAND_TOTAL_FILL
        c.font = GRAND_TOTAL_FONT
        c.alignment = CENTER if col_idx in (9, 13) else RIGHT
        c.border = THIN_BORDER

    ws[f"{COL_TOTAL_COST}{row}"].number_format = '#,##0" ₪"'
    ws[f"{COL_TOTAL_CHARGE}{row}"].number_format = '#,##0" ₪"'
    ws[f"{COL_PROFIT}{row}"].number_format = '#,##0" ₪"'
    ws[f"{COL_MARGIN}{row}"].number_format = "0.0%"
    ws.row_dimensions[row].height = 30


def build(budget_json_path, output_xlsx_path):
    with open(budget_json_path, encoding="utf-8") as f:
        data = json.load(f)

    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_NAME
    ws.sheet_view.rightToLeft = True

    # Column widths.
    for col_letter, width in COLUMN_WIDTHS.items():
        ws.column_dimensions[col_letter].width = width

    write_header_row(ws)

    # Main section: group items under the 6 canonicals.
    main_buckets = group_items_by_canonical(data.get("categories", []))

    cursor = 2
    main_first_data_row = None
    main_last_data_row = None

    for category in CANONICAL_CATEGORIES:
        items = main_buckets[category]
        if not items:
            continue
        write_category_marker(ws, cursor, category)
        cursor += 1
        for item in items:
            if main_first_data_row is None:
                main_first_data_row = cursor
            write_line_item(ws, cursor, item, category)
            main_last_data_row = cursor
            cursor += 1

    # Main subtotal, production fee, grand total.
    if main_first_data_row is not None:
        main_subtotal_row = cursor
        write_section_subtotal(
            ws, main_subtotal_row, 'סה"כ משני',
            main_first_data_row, main_last_data_row,
        )
        cursor += 1
        main_fee_row = cursor
        write_production_fee(ws, main_fee_row, main_subtotal_row)
        cursor += 1
        main_total_row = cursor
        write_grand_total(
            ws, main_total_row, main_subtotal_row, main_fee_row,
            label='סה"כ חשבונית',
        )
        cursor += 3  # blank rows before options

    # Options / conditional section.
    conditional_items = data.get("conditional_items", [])
    if conditional_items:
        # "אופציות" header row.
        ws.cell(row=cursor, column=8).value = "אופציות"
        for col_idx in range(1, 14):
            c = ws.cell(row=cursor, column=col_idx)
            c.fill = HEADER_FILL
            c.font = HEADER_FONT
            c.alignment = CENTER
            c.border = THIN_BORDER
        ws.row_dimensions[cursor].height = 28
        cursor += 1

        # Single כללי marker.
        write_category_marker(ws, cursor, "כללי")
        cursor += 1

        opt_first_data_row = cursor
        for item in conditional_items:
            write_line_item(ws, cursor, item, "כללי")
            cursor += 1
        opt_last_data_row = cursor - 1

        opt_subtotal_row = cursor
        write_section_subtotal(
            ws, opt_subtotal_row, 'סה"כ משני',
            opt_first_data_row, opt_last_data_row,
        )
        cursor += 1
        opt_fee_row = cursor
        write_production_fee(ws, opt_fee_row, opt_subtotal_row)
        cursor += 1
        opt_total_row = cursor
        write_grand_total(
            ws, opt_total_row, opt_subtotal_row, opt_fee_row,
            label='סה"כ חשבונית',
        )

    wb.save(output_xlsx_path)
    print(f"  written: {output_xlsx_path}")


if __name__ == "__main__":
    if len(sys.argv) != 3:
        print("usage: python build_budget_xlsx.py <budget.json> <output.xlsx>")
        sys.exit(1)
    build(sys.argv[1], sys.argv[2])
